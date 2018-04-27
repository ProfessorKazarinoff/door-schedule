
# coding: utf-8


from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill, colors

import os

class instructorObj():
    def __init__(self, name):
        self.name = name
        self.classes = []
        self.departments = []
        self.office = ''
        self.phone = ''
        self.email = ''
        self.year = ''
        self.quarter = ''

    def __str__(self):
        return self.name

    def print_schedule(self):
        print(self.name)
        print()
        for x in self.classes:
            print(x.course_number)
            print(x.course_name)
            print(x.day)
            print(x.start_time)
            print(x.end_time)
            print(x.building)
            print(x.room_number)
            print()

def insert_class_sec(wbsheetObj, course_number='CMET 235', building='AM', room='105', day='Tu', start_time='11:00 AM', end_time='2:30 PM', color='ADD8E6'):
    """
    function inserts a class section into a worksheet object
    """
    if building and room:
        building_and_room = "".join([building,room])
    else: building_and_room ="Bld Room#"

    day_dict = {'M':2,'Tu':3,'W':4,'MW': 4, 'Th':5,'F':6,'Sa':7,'Su':8}
    acronym_dict = {'Monday':2,'Tuesday':3,'Wednesday':4, 'Thursday':5,'Friday':6,'Saturday':7,'Sunday':8}
    if day in day_dict.keys():
        col = day_dict[day]
    else:
        col = day_dict['Su']
        color = "FF9999"

    if (not start_time) or (not end_time):
        start_time='8:00 AM'
        end_time='9:00 AM'
        color = "FF9999"

    lead_num = get_24h_dec_time(start_time)
    above_7 = lead_num - 7
    row = 7 + above_7*2

    lead_num = get_24h_dec_time(end_time)
    above_7 = lead_num - 7
    merge_end = 7 + above_7*2 - 1
    
    wbsheetObj.merge_cells(start_row=row, start_column=col, end_row=merge_end, end_column=col)
    wbsheetObj.cell(row=row, column=col).value = "\n".join([course_number, building_and_room])
    al=Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    wbsheetObj.cell(row=row, column=col).alignment = al
    fl=PatternFill(fill_type='solid', start_color=color)
    wbsheetObj.cell(row=row, column=col).fill = fl
    
    return  wbsheetObj

def get_24h_dec_time(time_str):
    """
    This function takes in a time string of the form '2:30 PM' and outputs a 24-hour decimal hour representation of 14.5.
    All minutes are rounded to half hour increments
    """
    if time_str:
        hour = time_str.split(':')[0]
        min_chunk = time_str.split(':')[1]
        min_sm_chunk = min_chunk.replace(" AM", "")
        minute = min_sm_chunk.replace(" PM","")
        if 'PM' in time_str and not hour == '12':
            hour = int(hour)+12
        elif 'PM' in time_str and hour == '12':
            hour = 12
        else:
            hour = int(hour)
        min_dec = float(minute)/60
        min_round = round(float(min_dec) * 2) / 2
        return hour+min_round
    else:
        return None

def insert_gen_info(wbsheetObj, quarter='spring', year='2018', instructor ='Peter Kazarinoff', department='CMET ENGR', email='peter.kazarinoff@pcc.edu', phone='971.722.8065'):
    """
    function builds the header part of an excel schedule 
    """
    wbsheetObj['D1'] = quarter.upper()
    wbsheetObj['D1'] = year
    wbsheetObj['C3'] = instructor
    wbsheetObj['C4'] = department

    wbsheetObj['F3'] = email
    wbsheetObj['F4'] = phone
    
    al=Alignment(horizontal='left', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=1)

    c3 = wbsheetObj['C3']
    c4 = wbsheetObj['C4']
    f3 = wbsheetObj['F3']
    f4 = wbsheetObj['F4']
    c3.alignment = al
    c4.alignment = al
    f3.alignment = al
    f4.alignment = al
    
    return wbsheetObj

def main():
    template_path = os.path.join(os.getcwd(),'templates','schedule_template.xlsx')
    wb = load_workbook(template_path)
    ws = wb['Sheet1']
    ws = insert_gen_info(ws)
    ws = insert_class_sec(ws)

    wkbk_path = os.path.join(os.getcwd(),'out','spreadsheet.xlsx')
    if not os.path.exists(os.path.join(os.getcwd(),'out')):
        os.mkdir('out')
    wb.save(wkbk_path)

if __name__ == "__main__":
    main()
    
    
    
